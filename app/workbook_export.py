from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from sqlalchemy import text
from sqlalchemy.orm import Session

MONTH_SHORT = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
MONTH_NAMES = [
    "JANUARY",
    "FEBRUARY",
    "MARCH",
    "APRIL",
    "MAY",
    "JUNE",
    "JULY",
    "AUGUST",
    "SEPTEMBER",
    "OCTOBER",
    "NOVEMBER",
    "DECEMBER",
]


def _text_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _sheet_by_exact(workbook, wanted: str):
    target = wanted.strip().lower()
    for name in workbook.sheetnames:
        if name.strip().lower() == target:
            return workbook[name]
    return None


def _sheet_by_prefix(workbook, prefix: str):
    target = prefix.strip().lower()
    for name in workbook.sheetnames:
        if name.strip().lower().startswith(target):
            return workbook[name]
    return None


def _ensure_sheet(workbook, name: str):
    existing = _sheet_by_exact(workbook, name)
    if existing is not None:
        return existing
    return workbook.create_sheet(title=name)


def _clear_sheet(sheet) -> None:
    if sheet.max_row > 0:
        sheet.delete_rows(1, sheet.max_row)


def _clear_range_values(sheet, *, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    if end_row < start_row:
        return
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            sheet.cell(row=row, column=col).value = None


def _resolve_template_sheet(workbook, kind: str):
    if kind == "get_data":
        return _sheet_by_prefix(workbook, "Get Data -")
    if kind == "customer_details":
        return _sheet_by_prefix(workbook, "Customer Details")
    if kind == "cvm":
        return _sheet_by_exact(workbook, " CVM") or _sheet_by_exact(workbook, "CVM")
    if kind == "database":
        return _sheet_by_exact(workbook, "Database")
    return None


def _load_workbook_template():
    try:
        from openpyxl import Workbook, load_workbook
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail="Workbook export requires openpyxl. Install dependencies and redeploy.",
        ) from exc

    template_path = Path("app/templates/planner_template.xlsm")
    if template_path.exists():
        workbook = load_workbook(template_path, keep_vba=True)
        return workbook, "xlsm", True

    workbook = Workbook()
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])
    return workbook, "xlsx", False


def _resolve_export_year(db: Session, year: int | None) -> int:
    if year is not None:
        return year
    settings = db.execute(
        text("SELECT calendar_year FROM calendar_settings WHERE id = 1")
    ).mappings().first()
    if settings and settings.get("calendar_year"):
        return int(settings["calendar_year"])
    return date.today().year


def _load_export_data(db: Session, export_year: int, territory_id: int | None) -> dict[str, Any]:
    params = {"territory_id": territory_id, "year": export_year}

    customer_rows = db.execute(
        text(
            """
            SELECT c.id,
                   COALESCE(t.name, '') AS territory,
                   COALESCE(c.cust_code, '') AS cust_code,
                   COALESCE(c.name, '') AS customer_name,
                   COALESCE(c.trade_name, '') AS trade_name,
                   COALESCE(c.group_name, '') AS group_name,
                   COALESCE(c.group_2_iws, '') AS group_2_iws,
                   COALESCE(c.iws_code, '') AS iws_code,
                   COALESCE(c.old_value, '') AS old_value,
                   COALESCE(c.old_name, '') AS old_name,
                   COALESCE(c.door_count, 0) AS door_count,
                   COALESCE(c.cvm_notes, '') AS cvm_notes,
                   COALESCE(
                     (
                       SELECT s.sort_bucket
                       FROM stores s
                       WHERE s.customer_id = c.id
                       ORDER BY s.id
                       LIMIT 1
                     ),
                     COALESCE(c.group_name, '')
                   ) AS sort_bucket
            FROM customers c
            LEFT JOIN territories t ON t.id = c.territory_id
            WHERE (:territory_id IS NULL OR c.territory_id = :territory_id)
            ORDER BY c.cust_code
            """
        ),
        params,
    ).mappings().all()

    store_rows = db.execute(
        text(
            """
            SELECT s.customer_id,
                   COALESCE(s.address_1, '') AS address_1,
                   COALESCE(s.address_2, '') AS address_2,
                   COALESCE(s.city, '') AS city,
                   COALESCE(s.state, '') AS state,
                   COALESCE(s.postcode, '') AS postcode,
                   COALESCE(s.country, '') AS country,
                   COALESCE(s.main_contact, '') AS main_contact,
                   COALESCE(s.owner_name, '') AS owner_name,
                   COALESCE(s.owner_phone, '') AS owner_phone,
                   COALESCE(s.owner_email, '') AS owner_email,
                   COALESCE(s.store_manager_name, '') AS store_manager_name,
                   COALESCE(s.store_phone, '') AS store_phone,
                   COALESCE(s.store_email, '') AS store_email,
                   COALESCE(s.market_manager_name, '') AS market_manager_name,
                   COALESCE(s.marketing_phone, '') AS marketing_phone,
                   COALESCE(s.marketing_email, '') AS marketing_email,
                   COALESCE(s.account_dept_name, '') AS account_dept_name,
                   COALESCE(s.accounting_phone, '') AS accounting_phone,
                   COALESCE(s.accounting_email, '') AS accounting_email,
                   COALESCE(s.notes, '') AS notes
            FROM stores s
            JOIN customers c ON c.id = s.customer_id
            WHERE (:territory_id IS NULL OR c.territory_id = :territory_id)
            ORDER BY c.cust_code, s.id
            """
        ),
        params,
    ).mappings().all()

    cvm_rows = db.execute(
        text(
            """
            SELECT e.customer_id, e.month, e.planned_date, e.completed_manual
            FROM cvm_month_entries e
            JOIN customers c ON c.id = e.customer_id
            WHERE e.year = :year
              AND (:territory_id IS NULL OR c.territory_id = :territory_id)
            ORDER BY e.customer_id, e.month
            """
        ),
        params,
    ).mappings().all()

    last_visit_rows = db.execute(
        text(
            """
            SELECT e.customer_id, MAX(e.planned_date) AS last_visit
            FROM cvm_month_entries e
            JOIN customers c ON c.id = e.customer_id
            WHERE e.completed_manual = TRUE
              AND e.planned_date IS NOT NULL
              AND (:territory_id IS NULL OR c.territory_id = :territory_id)
            GROUP BY e.customer_id
            """
        ),
        {"territory_id": territory_id},
    ).mappings().all()

    product_rows = db.execute(
        text(
            """
            SELECT p.customer_id,
                   COALESCE(p.product_name, '') AS product_name,
                   p.last_visit,
                   COALESCE(p.action, '') AS action,
                   COALESCE(p.status, '') AS status,
                   COALESCE(p.next_action, '') AS next_action,
                   p.last_contact,
                   COALESCE(p.notes, '') AS notes,
                   p.updated_at
            FROM products p
            JOIN customers c ON c.id = p.customer_id
            WHERE (:territory_id IS NULL OR c.territory_id = :territory_id)
            ORDER BY c.cust_code, p.product_name, p.updated_at DESC, p.id DESC
            """
        ),
        {"territory_id": territory_id},
    ).mappings().all()

    stores_by_customer: dict[int, list[dict[str, Any]]] = {}
    for row in store_rows:
        stores_by_customer.setdefault(int(row["customer_id"]), []).append(dict(row))

    month_entries: dict[int, dict[int, dict[str, Any]]] = {}
    for row in cvm_rows:
        customer_id = int(row["customer_id"])
        month = int(row["month"])
        month_entries.setdefault(customer_id, {})[month] = {
            "planned_date": row["planned_date"],
            "completed_manual": bool(row["completed_manual"]),
        }

    last_visits = {int(row["customer_id"]): row["last_visit"] for row in last_visit_rows}

    products_by_customer: dict[int, dict[str, dict[str, Any]]] = {}
    product_names: list[str] = []
    seen_product_names: set[str] = set()
    for row in product_rows:
        customer_id = int(row["customer_id"])
        product_name = _text_value(row["product_name"])
        if not product_name:
            continue
        row_map = products_by_customer.setdefault(customer_id, {})
        if product_name in row_map:
            continue
        row_map[product_name] = dict(row)
        if product_name not in seen_product_names:
            seen_product_names.add(product_name)
            product_names.append(product_name)

    product_names.sort(key=lambda value: value.lower())

    return {
        "customers": [dict(row) for row in customer_rows],
        "stores_by_customer": stores_by_customer,
        "month_entries": month_entries,
        "last_visits": last_visits,
        "products_by_customer": products_by_customer,
        "product_names": product_names,
    }


def _write_get_data_sheet(sheet, customers: list[dict[str, Any]], *, preserve_template: bool) -> None:
    if preserve_template:
        header_row = 1
        headers = [
            "Territory",
            "Group",
            "Group 2 IWS",
            "IWS Codes",
            "Customer Number",
            "Customer Name",
            "OLD Value",
            "Old Name",
        ]
        for idx, header in enumerate(headers, start=1):
            sheet.cell(row=header_row, column=idx).value = header
        start_row = 2
        max_row = max(sheet.max_row, start_row + len(customers) - 1)
        _clear_range_values(sheet, start_row=start_row, end_row=max_row, start_col=1, end_col=8)
        for idx, customer in enumerate(customers):
            row = start_row + idx
            sheet.cell(row=row, column=1).value = customer.get("territory") or ""
            sheet.cell(row=row, column=2).value = customer.get("group_name") or ""
            sheet.cell(row=row, column=3).value = customer.get("group_2_iws") or ""
            sheet.cell(row=row, column=4).value = customer.get("iws_code") or ""
            sheet.cell(row=row, column=5).value = customer.get("cust_code") or ""
            sheet.cell(row=row, column=6).value = customer.get("customer_name") or ""
            sheet.cell(row=row, column=7).value = customer.get("old_value") or ""
            sheet.cell(row=row, column=8).value = customer.get("old_name") or ""
        return

    _clear_sheet(sheet)
    sheet.append(
        [
            "Territory",
            "Group",
            "Group 2 IWS",
            "IWS Codes",
            "Customer Number",
            "Customer Name",
            "OLD Value",
            "Old Name",
        ]
    )
    for customer in customers:
        sheet.append(
            [
                customer.get("territory") or "",
                customer.get("group_name") or "",
                customer.get("group_2_iws") or "",
                customer.get("iws_code") or "",
                customer.get("cust_code") or "",
                customer.get("customer_name") or "",
                customer.get("old_value") or "",
                customer.get("old_name") or "",
            ]
        )


def _write_customer_details_sheet(
    sheet,
    customers: list[dict[str, Any]],
    stores_by_customer: dict[int, list[dict[str, Any]]],
    *,
    preserve_template: bool,
) -> None:
    output_rows: list[list[Any]] = []
    for customer in customers:
        cust_code = customer.get("cust_code") or ""
        customer_name = customer.get("customer_name") or ""
        customer_object = f"{cust_code} | {customer_name}".strip(" |")
        customer_id = int(customer["id"])
        store_rows = stores_by_customer.get(customer_id) or [{}]
        for store in store_rows:
            output_rows.append(
                [
                    cust_code,
                    customer_object,
                    customer_object,
                    customer.get("territory") or "",
                    "",
                    store.get("address_1", ""),
                    store.get("address_2", ""),
                    store.get("city", ""),
                    store.get("state", ""),
                    store.get("postcode", ""),
                    store.get("country", ""),
                    store.get("main_contact", ""),
                    customer.get("trade_name") or "",
                    store.get("owner_name", ""),
                    store.get("owner_phone", ""),
                    store.get("owner_email", ""),
                    store.get("store_manager_name", ""),
                    store.get("store_phone", ""),
                    store.get("store_email", ""),
                    store.get("market_manager_name", ""),
                    store.get("marketing_phone", ""),
                    store.get("marketing_email", ""),
                    store.get("account_dept_name", ""),
                    store.get("accounting_phone", ""),
                    store.get("accounting_email", ""),
                ]
            )

    if preserve_template:
        start_row = 3
        max_row = max(sheet.max_row, start_row + len(output_rows) - 1)
        _clear_range_values(sheet, start_row=start_row, end_row=max_row, start_col=1, end_col=25)
        for idx, values in enumerate(output_rows):
            row = start_row + idx
            for col, value in enumerate(values, start=1):
                sheet.cell(row=row, column=col).value = value
        return

    _clear_sheet(sheet)
    sheet.append([])
    sheet.append(
        [
            "Custom",
            "Customer Object",
            "Customer Object (BillTo)",
            "Customer Territory A Name",
            "Customer Territory B Name",
            "STORE ADDRESS 1",
            "STORE ADDRESS 2",
            "SUBURB",
            "STATE",
            "POSTCODE",
            "COUNTRY",
            "MAIN CONTACT",
            "OWNER NAME",
            "OWNER PHONE",
            "OWNER EMAIL",
            "STORE MANAGER NAME",
            "STORE PHONE",
            "STORE EMAIL",
            "MARKET MANAGER NAME",
            "MARKETING PHONE",
            "MARKETING EMAIL",
            "ACCOUNTS NAME",
            "ACCOUNTING PHONE",
            "ACCOUNTING EMAIL",
        ]
    )
    for values in output_rows:
        sheet.append(values)


def _write_cvm_sheet(
    sheet,
    customers: list[dict[str, Any]],
    month_entries: dict[int, dict[int, dict[str, Any]]],
    *,
    preserve_template: bool,
) -> None:
    if preserve_template:
        start_row = 4
        last_row = max(sheet.max_row, start_row + len(customers) - 1)
        manual_columns = [1, 4, 7] + list(range(11, 35))
        for row in range(start_row, last_row + 1):
            for col in manual_columns:
                sheet.cell(row=row, column=col).value = None

        for idx, customer in enumerate(customers):
            row = start_row + idx
            customer_id = int(customer["id"])
            customer_months = month_entries.get(customer_id, {})
            sheet.cell(row=row, column=1).value = customer.get("door_count") or 0
            sheet.cell(row=row, column=4).value = customer.get("sort_bucket") or ""
            sheet.cell(row=row, column=7).value = customer.get("cvm_notes") or ""
            for month_index in range(1, 13):
                cell = customer_months.get(month_index, {})
                planned_col = 11 + (month_index - 1) * 2
                completed_col = planned_col + 1
                sheet.cell(row=row, column=planned_col).value = cell.get("planned_date")
                if "completed_manual" in cell:
                    sheet.cell(row=row, column=completed_col).value = bool(cell.get("completed_manual"))
                else:
                    sheet.cell(row=row, column=completed_col).value = False
        return

    _clear_sheet(sheet)
    sheet.append([])
    sheet.append([])
    headers = [
        "Door Count",
        "TERRITORY",
        "Cust Code",
        "SORT",
        "Customer Name",
        "TRADE NAME",
        "Notes/Comments",
        "Date of last completed visit",
        "Count of Planned",
        "Count of Visit",
    ]
    for month_short in MONTH_SHORT:
        headers.extend([f"PLANNED {month_short}", f"COMPLETED {month_short}"])
    sheet.append(headers)

    for customer in customers:
        customer_id = int(customer["id"])
        customer_months = month_entries.get(customer_id, {})
        planned_total = sum(1 for cell in customer_months.values() if cell.get("planned_date"))
        completed_total = sum(1 for cell in customer_months.values() if cell.get("completed_manual"))
        completed_dates = [
            cell.get("planned_date")
            for cell in customer_months.values()
            if cell.get("completed_manual") and cell.get("planned_date") is not None
        ]
        last_completed = max(completed_dates) if completed_dates else None

        row: list[Any] = [
            customer.get("door_count") or 0,
            customer.get("territory") or "",
            customer.get("cust_code") or "",
            customer.get("sort_bucket") or "",
            customer.get("customer_name") or "",
            customer.get("trade_name") or "",
            customer.get("cvm_notes") or "",
            last_completed,
            planned_total,
            completed_total,
        ]
        for month_index in range(1, 13):
            cell = customer_months.get(month_index, {})
            row.append(cell.get("planned_date"))
            if "completed_manual" in cell:
                row.append(bool(cell.get("completed_manual")))
            else:
                row.append("")
        sheet.append(row)


def _write_month_sheets(workbook, export_year: int) -> None:
    for month_name in MONTH_NAMES:
        month_sheet = _ensure_sheet(workbook, month_name)
        month_sheet["R4"] = export_year


def _write_database_sheet(
    sheet,
    customers: list[dict[str, Any]],
    products_by_customer: dict[int, dict[str, dict[str, Any]]],
    product_names: list[str],
    last_visits: dict[int, Any],
    *,
    preserve_template: bool,
) -> None:
    if preserve_template:
        action_group_start = 28
        group_width = 5
        template_group_count = max((sheet.max_column - action_group_start + 1) // group_width, 0)
        export_product_names = product_names[:template_group_count] if template_group_count else []
        end_col = sheet.max_column

        # Keep Database totals formulas (cols 26-27) intact while replacing metadata/action blocks.
        _clear_range_values(sheet, start_row=3, end_row=4, start_col=action_group_start, end_col=end_col)
        for idx, product_name in enumerate(export_product_names):
            col = action_group_start + idx * group_width
            sheet.cell(row=3, column=col).value = product_name
            sheet.cell(row=4, column=col).value = "ACTION"
            sheet.cell(row=4, column=col + 1).value = "STATUS"
            sheet.cell(row=4, column=col + 2).value = "NEXT ACTION"
            sheet.cell(row=4, column=col + 3).value = "LAST CONTACT \n[Enter dates only]"
            sheet.cell(row=4, column=col + 4).value = "NOTES"

        start_row = 5
        max_row = max(sheet.max_row, start_row + len(customers) - 1)
        _clear_range_values(sheet, start_row=start_row, end_row=max_row, start_col=20, end_col=25)
        _clear_range_values(
            sheet,
            start_row=start_row,
            end_row=max_row,
            start_col=action_group_start,
            end_col=end_col,
        )

        for idx, customer in enumerate(customers):
            row = start_row + idx
            customer_id = int(customer["id"])
            sheet.cell(row=row, column=20).value = customer.get("door_count") or 0
            sheet.cell(row=row, column=21).value = customer.get("territory") or ""
            sheet.cell(row=row, column=22).value = customer.get("cust_code") or ""
            sheet.cell(row=row, column=23).value = customer.get("customer_name") or ""
            sheet.cell(row=row, column=24).value = customer.get("trade_name") or ""
            sheet.cell(row=row, column=25).value = last_visits.get(customer_id)
            customer_products = products_by_customer.get(customer_id, {})
            for prod_idx, product_name in enumerate(export_product_names):
                product = customer_products.get(product_name)
                if not product:
                    continue
                col = action_group_start + prod_idx * group_width
                sheet.cell(row=row, column=col).value = product.get("action") or ""
                sheet.cell(row=row, column=col + 1).value = product.get("status") or ""
                sheet.cell(row=row, column=col + 2).value = product.get("next_action") or ""
                sheet.cell(row=row, column=col + 3).value = product.get("last_contact")
                sheet.cell(row=row, column=col + 4).value = product.get("notes") or ""
        return

    _clear_sheet(sheet)

    metadata_headers = {
        21: ("TERRITORY", "TERRITORY"),
        22: ("CUST CODE", "CUST CODE"),
        23: ("CUSTOMER NAME", "CUSTOMER NAME"),
        24: ("TRADE NAME", "TRADE NAME"),
        25: ("LAST VISIT", "LAST VISIT"),
    }
    for col_idx, (product_label, field_label) in metadata_headers.items():
        sheet.cell(row=3, column=col_idx).value = product_label
        sheet.cell(row=4, column=col_idx).value = field_label

    action_group_start = 28
    for idx, product_name in enumerate(product_names):
        col = action_group_start + idx * 5
        sheet.cell(row=3, column=col).value = product_name
        sheet.cell(row=4, column=col).value = "ACTION"
        sheet.cell(row=4, column=col + 1).value = "STATUS"
        sheet.cell(row=4, column=col + 2).value = "NEXT ACTION"
        sheet.cell(row=4, column=col + 3).value = "LAST CONTACT"
        sheet.cell(row=4, column=col + 4).value = "NOTES"

    row_index = 5
    for customer in customers:
        customer_id = int(customer["id"])
        sheet.cell(row=row_index, column=21).value = customer.get("territory") or ""
        sheet.cell(row=row_index, column=22).value = customer.get("cust_code") or ""
        sheet.cell(row=row_index, column=23).value = customer.get("customer_name") or ""
        sheet.cell(row=row_index, column=24).value = customer.get("trade_name") or ""
        sheet.cell(row=row_index, column=25).value = last_visits.get(customer_id)

        customer_products = products_by_customer.get(customer_id, {})
        for idx, product_name in enumerate(product_names):
            product = customer_products.get(product_name)
            if not product:
                continue
            col = action_group_start + idx * 5
            sheet.cell(row=row_index, column=col).value = product.get("action") or ""
            sheet.cell(row=row_index, column=col + 1).value = product.get("status") or ""
            sheet.cell(row=row_index, column=col + 2).value = product.get("next_action") or ""
            sheet.cell(row=row_index, column=col + 3).value = product.get("last_contact")
            sheet.cell(row=row_index, column=col + 4).value = product.get("notes") or ""

        row_index += 1


def export_planner_workbook(db: Session, *, year: int | None = None, territory_id: int | None = None) -> dict[str, Any]:
    workbook, extension, template_mode = _load_workbook_template()
    export_year = _resolve_export_year(db, year)
    data = _load_export_data(db, export_year=export_year, territory_id=territory_id)

    get_data_sheet = _resolve_template_sheet(workbook, "get_data") or _ensure_sheet(workbook, "Get Data -Sample")
    details_sheet = _resolve_template_sheet(workbook, "customer_details") or _ensure_sheet(workbook, "Customer Details ")
    cvm_sheet = _resolve_template_sheet(workbook, "cvm") or _ensure_sheet(workbook, "CVM")
    database_sheet = _resolve_template_sheet(workbook, "database") or _ensure_sheet(workbook, "Database")

    _write_get_data_sheet(get_data_sheet, data["customers"], preserve_template=template_mode)
    _write_customer_details_sheet(
        details_sheet,
        data["customers"],
        data["stores_by_customer"],
        preserve_template=template_mode,
    )
    _write_cvm_sheet(
        cvm_sheet,
        data["customers"],
        data["month_entries"],
        preserve_template=template_mode,
    )
    _write_month_sheets(workbook, export_year)
    _write_database_sheet(
        database_sheet,
        data["customers"],
        data["products_by_customer"],
        data["product_names"],
        data["last_visits"],
        preserve_template=template_mode,
    )

    buffer = BytesIO()
    workbook.save(buffer)
    return {
        "content": buffer.getvalue(),
        "extension": extension,
        "year": export_year,
    }
