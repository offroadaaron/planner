from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from fastapi import HTTPException
from sqlalchemy import text
from sqlalchemy.orm import Session

VALID_UPSERT_POLICIES = {"merge", "create_only", "overwrite"}
VALID_VALIDATION_MODES = {"standard", "strict"}
VALID_DUPLICATE_POLICIES = {"last_wins", "first_wins", "error"}
ROW_ISSUE_LIMIT = 300
MONTH_SHORT = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
MONTH_NAMES = [
    "JANUARY",
    "FEBRUARY",
    "MARCH",
    "APRIL",
    "MAY",
    "JUNE",
    "JULY",
    "AUGUST",
    "SEPTEMBER",
    "OCTOBER",
    "NOVEMBER",
    "DECEMBER",
]


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    text_value = str(value).replace("\xa0", " ").strip()
    return text_value


def _clean_code(value: Any) -> str:
    raw = _clean_text(value)
    if not raw:
        return ""
    if raw in {"0", "0.0"}:
        return ""
    if raw.endswith(".0") and raw.replace(".", "", 1).isdigit():
        return raw[:-2]
    return raw


def _normalize_header_label(value: Any) -> str:
    raw = _clean_text(value).upper().replace("_", " ").replace("-", " ").replace("/", " ")
    cleaned = "".join(ch if (ch.isalnum() or ch == " ") else " " for ch in raw)
    return " ".join(cleaned.split())


def _find_header_column(headers: list[Any], candidates: set[str]) -> int | None:
    normalized_candidates = {_normalize_header_label(c) for c in candidates if _clean_text(c)}
    if not normalized_candidates:
        return None
    for idx, header in enumerate(headers, start=1):
        if _normalize_header_label(header) in normalized_candidates:
            return idx
    return None


def _column_value(row: tuple[Any, ...] | list[Any], column_index: int | None) -> Any:
    if column_index is None or column_index <= 0:
        return None
    if column_index > len(row):
        return None
    return row[column_index - 1]


def _is_effectively_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, bool):
        return not value
    if isinstance(value, (int, float)):
        return value == 0
    return _clean_text(value) == ""


def _has_meaningful_values(values: list[Any]) -> bool:
    return any(not _is_effectively_empty(v) for v in values)


def _extract_name(value: Any) -> str:
    raw = _clean_text(value)
    if not raw:
        return ""
    if "|" in raw:
        parts = [p.strip() for p in raw.split("|") if p.strip()]
        if len(parts) >= 2:
            return parts[-1]
    return raw


def _to_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    raw = _clean_text(value)
    if not raw:
        return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value > 0
    raw = _clean_text(value).lower()
    if not raw:
        return False
    return raw in {"true", "yes", "y", "1", "done", "completed", "x"}


def _to_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return None

    raw = _clean_text(value)
    if not raw:
        return None
    if raw.endswith(".0") and raw.replace(".", "", 1).isdigit():
        raw = raw[:-2]
    if raw.lstrip("-").isdigit():
        return int(raw)
    return None


def _sheet_by_prefix(workbook, prefix: str):
    target = prefix.strip().lower()
    for name in workbook.sheetnames:
        if name.strip().lower().startswith(target):
            return workbook[name]
    return None


def _sheet_by_exact(workbook, wanted: str):
    target = wanted.strip().lower()
    for name in workbook.sheetnames:
        if name.strip().lower() == target:
            return workbook[name]
    return None


def _resolve_calendar_year(workbook) -> int | None:
    def _year_from_value(value: Any) -> int | None:
        if isinstance(value, (int, float)):
            year_value = int(value)
            return year_value if 2000 <= year_value <= 2100 else None
        raw = _clean_text(value)
        if raw.isdigit():
            year_value = int(raw)
            return year_value if 2000 <= year_value <= 2100 else None
        return None

    for month_name in MONTH_NAMES:
        month_sheet = _sheet_by_exact(workbook, month_name)
        if month_sheet is None:
            continue
        year_value = _year_from_value(month_sheet["R4"].value)
        if year_value is not None:
            return year_value

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        year_value = _year_from_value(sheet["R4"].value)
        if year_value is not None:
            return year_value

    return None


def _resolve_cvm_column_map(cvm_sheet) -> dict[str, Any]:
    max_col = cvm_sheet.max_column
    header_row = [_clean_text(cvm_sheet.cell(row=3, column=i).value) for i in range(1, max_col + 1)]

    resolved: dict[str, Any] = {
        "door_count_col": _find_header_column(header_row, {"DOOR COUNT"}) or 1,
        "territory_col": _find_header_column(header_row, {"TERRITORY"}) or 2,
        "cust_code_col": _find_header_column(header_row, {"CUST CODE", "CUSTOMER CODE", "CUSTOMER NUMBER"}) or 3,
        "sort_col": _find_header_column(header_row, {"SORT"}) or 4,
        "customer_name_col": _find_header_column(header_row, {"CUSTOMER NAME"}) or 5,
        "trade_name_col": _find_header_column(header_row, {"TRADE NAME"}) or 6,
        "notes_col": _find_header_column(header_row, {"NOTES COMMENTS", "NOTES", "COMMENTS"}) or 7,
        "month_cols": {},
    }

    month_cols: dict[int, dict[str, int]] = {}
    for col_idx, label in enumerate(header_row, start=1):
        normalized = _normalize_header_label(label)
        for month_idx, month_short in enumerate(MONTH_SHORT, start=1):
            if normalized == f"PLANNED {month_short}":
                month_cols.setdefault(month_idx, {})["planned_col"] = col_idx
            elif normalized in {f"COMPLETED {month_short}", f"COMPLETE {month_short}", f"DONE {month_short}"}:
                month_cols.setdefault(month_idx, {})["completed_col"] = col_idx

    # Legacy fallback if month headers are missing or shifted.
    for month_idx in range(1, 13):
        legacy_planned_col = 11 + (month_idx - 1) * 2
        legacy_completed_col = legacy_planned_col + 1
        cols = month_cols.setdefault(month_idx, {})
        if "planned_col" not in cols and legacy_planned_col <= max_col:
            cols["planned_col"] = legacy_planned_col
        if "completed_col" not in cols and legacy_completed_col <= max_col:
            cols["completed_col"] = legacy_completed_col

    resolved["month_cols"] = month_cols
    return resolved


def _find_database_column(field_labels: list[str], product_labels: list[str], candidates: set[str]) -> int | None:
    found = _find_header_column(field_labels, candidates)
    if found is not None:
        return found
    return _find_header_column(product_labels, candidates)


def _cell_value(sheet, row_idx: int, col_idx: int | None) -> Any:
    if col_idx is None or col_idx <= 0:
        return None
    return sheet.cell(row=row_idx, column=col_idx).value


def _resolve_database_column_map(database_sheet) -> dict[str, int]:
    max_col = database_sheet.max_column
    product_labels = [_clean_text(database_sheet.cell(row=3, column=i).value) for i in range(1, max_col + 1)]
    field_labels = [_clean_text(database_sheet.cell(row=4, column=i).value) for i in range(1, max_col + 1)]

    return {
        "territory_col": _find_database_column(field_labels, product_labels, {"TERRITORY"}) or 21,
        "cust_code_col": _find_database_column(
            field_labels,
            product_labels,
            {"CUST CODE", "CUSTOMER CODE", "CUSTOMER NUMBER"},
        )
        or 22,
        "customer_name_col": _find_database_column(field_labels, product_labels, {"CUSTOMER NAME"}) or 23,
        "trade_name_col": _find_database_column(field_labels, product_labels, {"TRADE NAME"}) or 24,
        "last_visit_col": _find_database_column(
            field_labels,
            product_labels,
            {"LAST VISIT", "DATE OF LAST COMPLETED VISIT", "DATE OF LAST VISIT"},
        )
        or 25,
    }


def _normalize_upsert_policy(upsert_policy: str) -> str:
    policy = _clean_text(upsert_policy).lower() or "merge"
    if policy not in VALID_UPSERT_POLICIES:
        allowed = ", ".join(sorted(VALID_UPSERT_POLICIES))
        raise HTTPException(status_code=400, detail=f"Invalid upsert policy '{upsert_policy}'. Allowed: {allowed}.")
    return policy


def _normalize_validation_mode(validation_mode: str) -> str:
    mode = _clean_text(validation_mode).lower() or "standard"
    if mode not in VALID_VALIDATION_MODES:
        allowed = ", ".join(sorted(VALID_VALIDATION_MODES))
        raise HTTPException(status_code=400, detail=f"Invalid validation mode '{validation_mode}'. Allowed: {allowed}.")
    return mode


def _normalize_duplicate_policy(duplicate_policy: str) -> str:
    policy = _clean_text(duplicate_policy).lower() or "last_wins"
    if policy not in VALID_DUPLICATE_POLICIES:
        allowed = ", ".join(sorted(VALID_DUPLICATE_POLICIES))
        raise HTTPException(
            status_code=400,
            detail=f"Invalid duplicate policy '{duplicate_policy}'. Allowed: {allowed}.",
        )
    return policy


def _record_issue(
    summary: dict[str, Any],
    *,
    level: str,
    sheet: str,
    row: int | None,
    message: str,
) -> None:
    issue = {
        "level": level,
        "sheet": sheet,
        "row": row,
        "message": message,
    }
    issues = summary.setdefault("row_issues", [])
    issue_limit = int(summary.get("row_issue_limit", ROW_ISSUE_LIMIT))
    if len(issues) < issue_limit:
        issues.append(issue)
    else:
        summary["row_issues_truncated"] = int(summary.get("row_issues_truncated", 0)) + 1

    if level == "error":
        summary["error_count"] = int(summary.get("error_count", 0)) + 1
    elif level == "warning":
        summary["warning_count"] = int(summary.get("warning_count", 0)) + 1


def _validation_level(summary: dict[str, Any]) -> str:
    return "error" if str(summary.get("validation_mode", "standard")) == "strict" else "warning"


def _add_blocker(summary: dict[str, Any], message: str) -> None:
    blockers = summary.setdefault("blockers", [])
    if message not in blockers:
        blockers.append(message)


def _is_row_populated(row: tuple[Any, ...] | list[Any]) -> bool:
    return any(_clean_text(v) for v in row)


def _register_duplicate(
    summary: dict[str, Any],
    seen: dict[str, int],
    *,
    key: str,
    sheet: str,
    row: int,
    label: str,
) -> bool:
    first_row = seen.get(key)
    if first_row is None:
        seen[key] = row
        return True

    duplicate_policy = str(summary.get("duplicate_policy", "last_wins"))
    base_message = f"Duplicate {label} key '{key}' (first seen at row {first_row})."

    if duplicate_policy == "last_wins":
        _record_issue(
            summary,
            level="warning",
            sheet=sheet,
            row=row,
            message=f"{base_message} Last row wins.",
        )
        return True

    if duplicate_policy == "first_wins":
        summary["duplicate_rows_skipped"] = int(summary.get("duplicate_rows_skipped", 0)) + 1
        _record_issue(
            summary,
            level="warning",
            sheet=sheet,
            row=row,
            message=f"{base_message} Row skipped (first row kept).",
        )
        return False

    summary["duplicate_rows_skipped"] = int(summary.get("duplicate_rows_skipped", 0)) + 1
    _record_issue(
        summary,
        level="error",
        sheet=sheet,
        row=row,
        message=f"{base_message} Row skipped (duplicate policy = error).",
    )
    _add_blocker(summary, "Duplicate key errors were found with duplicate policy set to 'error'.")
    return False


def _add_warning(summary: dict[str, Any], message: str, *, sheet: str | None = None, row: int | None = None) -> None:
    summary.setdefault("warnings", []).append(message)
    if sheet:
        _record_issue(summary, level="warning", sheet=sheet, row=row, message=message)
    else:
        summary["warning_count"] = int(summary.get("warning_count", 0)) + 1


def _to_date_with_issue(
    value: Any,
    *,
    summary: dict[str, Any],
    sheet: str,
    row: int,
    field: str,
) -> date | None:
    if value is None:
        return None
    parsed = _to_date(value)
    if parsed is not None:
        return parsed

    raw = _clean_text(value)
    if raw:
        level = _validation_level(summary)
        _record_issue(
            summary,
            level=level,
            sheet=sheet,
            row=row,
            message=f"Invalid date '{raw}' in {field}; value ignored.",
        )
    return None


def _get_or_create_territory(db: Session, territory_name: str, cache: dict[str, int], summary: dict[str, Any]) -> int | None:
    name = _clean_text(territory_name)
    if not name:
        return None
    if name in cache:
        return cache[name]

    existing = db.execute(
        text("SELECT id FROM territories WHERE name = :name"),
        {"name": name},
    ).mappings().first()
    if existing is not None:
        cache[name] = int(existing["id"])
        return cache[name]

    territory_id = db.execute(
        text("INSERT INTO territories (name) VALUES (:name) RETURNING id"),
        {"name": name},
    ).scalar_one()
    cache[name] = int(territory_id)
    summary["territories_created"] += 1
    return cache[name]


def _upsert_customer(
    db: Session,
    summary: dict[str, Any],
    cust_code: str,
    customer_name: str,
    territory_id: int | None,
    upsert_policy: str,
    *,
    trade_name: str = "",
    group_name: str = "",
    group_2_iws: str = "",
    iws_code: str = "",
    old_value: str = "",
    old_name: str = "",
    door_count: int | None = None,
    cvm_notes: str = "",
) -> int:
    code = _clean_code(cust_code)
    name = _clean_text(customer_name)
    if not code:
        raise HTTPException(status_code=400, detail="Workbook row is missing a customer code")

    existing = db.execute(
        text("SELECT id FROM customers WHERE cust_code = :cust_code"),
        {"cust_code": code},
    ).mappings().first()

    if existing is None:
        customer_id = db.execute(
            text(
                """
                INSERT INTO customers (
                  cust_code, name, trade_name, territory_id,
                  group_name, group_2_iws, iws_code,
                  old_value, old_name, door_count, cvm_notes, created_at
                )
                VALUES (
                  :cust_code, :name, NULLIF(:trade_name, ''), :territory_id,
                  NULLIF(:group_name, ''), NULLIF(:group_2_iws, ''), NULLIF(:iws_code, ''),
                  NULLIF(:old_value, ''), NULLIF(:old_name, ''), :door_count, NULLIF(:cvm_notes, ''), NOW()
                )
                RETURNING id
                """
            ),
            {
                "cust_code": code,
                "name": name or f"Customer {code}",
                "trade_name": _clean_text(trade_name),
                "territory_id": territory_id,
                "group_name": _clean_text(group_name),
                "group_2_iws": _clean_text(group_2_iws),
                "iws_code": _clean_text(iws_code),
                "old_value": _clean_text(old_value),
                "old_name": _clean_text(old_name),
                "door_count": door_count,
                "cvm_notes": _clean_text(cvm_notes),
            },
        ).scalar_one()
        summary["customers_created"] += 1
        return int(customer_id)

    customer_id = int(existing["id"])
    if upsert_policy == "create_only":
        summary["customers_skipped_existing"] += 1
        return customer_id

    if upsert_policy == "overwrite":
        db.execute(
            text(
                """
                UPDATE customers
                SET
                  name = CASE WHEN NULLIF(:name, '') IS NULL THEN name ELSE :name END,
                  trade_name = NULLIF(:trade_name, ''),
                  territory_id = :territory_id,
                  group_name = NULLIF(:group_name, ''),
                  group_2_iws = NULLIF(:group_2_iws, ''),
                  iws_code = NULLIF(:iws_code, ''),
                  old_value = NULLIF(:old_value, ''),
                  old_name = NULLIF(:old_name, ''),
                  door_count = COALESCE(:door_count, door_count),
                  cvm_notes = NULLIF(:cvm_notes, '')
                WHERE id = :customer_id
                """
            ),
            {
                "customer_id": customer_id,
                "name": name,
                "trade_name": _clean_text(trade_name),
                "territory_id": territory_id,
                "group_name": _clean_text(group_name),
                "group_2_iws": _clean_text(group_2_iws),
                "iws_code": _clean_text(iws_code),
                "old_value": _clean_text(old_value),
                "old_name": _clean_text(old_name),
                "door_count": door_count,
                "cvm_notes": _clean_text(cvm_notes),
            },
        )
        summary["customers_updated"] += 1
        return customer_id

    db.execute(
        text(
            """
            UPDATE customers
            SET
              name = COALESCE(NULLIF(:name, ''), name),
              trade_name = COALESCE(NULLIF(:trade_name, ''), trade_name),
              territory_id = COALESCE(:territory_id, territory_id),
              group_name = COALESCE(NULLIF(:group_name, ''), group_name),
              group_2_iws = COALESCE(NULLIF(:group_2_iws, ''), group_2_iws),
              iws_code = COALESCE(NULLIF(:iws_code, ''), iws_code),
              old_value = COALESCE(NULLIF(:old_value, ''), old_value),
              old_name = COALESCE(NULLIF(:old_name, ''), old_name),
              door_count = COALESCE(:door_count, door_count),
              cvm_notes = COALESCE(NULLIF(:cvm_notes, ''), cvm_notes)
            WHERE id = :customer_id
            """
        ),
        {
            "customer_id": customer_id,
            "name": name,
            "trade_name": _clean_text(trade_name),
            "territory_id": territory_id,
            "group_name": _clean_text(group_name),
            "group_2_iws": _clean_text(group_2_iws),
            "iws_code": _clean_text(iws_code),
            "old_value": _clean_text(old_value),
            "old_name": _clean_text(old_name),
            "door_count": door_count,
            "cvm_notes": _clean_text(cvm_notes),
        },
    )
    summary["customers_updated"] += 1
    return customer_id


def _upsert_store(
    db: Session,
    summary: dict[str, Any],
    customer_id: int,
    payload: dict[str, Any],
    upsert_policy: str,
) -> None:
    lookup = db.execute(
        text(
            """
            SELECT id
            FROM stores
            WHERE customer_id = :customer_id
              AND COALESCE(address_1, '') = :address_1
              AND COALESCE(city, '') = :city
              AND COALESCE(state, '') = :state
            ORDER BY id
            LIMIT 1
            """
        ),
        {
            "customer_id": customer_id,
            "address_1": _clean_text(payload.get("address_1")),
            "city": _clean_text(payload.get("city")),
            "state": _clean_text(payload.get("state")),
        },
    ).mappings().first()

    params = {
        "customer_id": customer_id,
        "address_1": _clean_text(payload.get("address_1")),
        "address_2": _clean_text(payload.get("address_2")),
        "city": _clean_text(payload.get("city")),
        "state": _clean_text(payload.get("state")),
        "postcode": _clean_text(payload.get("postcode")),
        "country": _clean_text(payload.get("country")),
        "main_contact": _clean_text(payload.get("main_contact")),
        "owner_name": _clean_text(payload.get("owner_name")),
        "owner_phone": _clean_text(payload.get("owner_phone")),
        "owner_email": _clean_text(payload.get("owner_email")),
        "store_manager_name": _clean_text(payload.get("store_manager_name")),
        "store_phone": _clean_text(payload.get("store_phone")),
        "store_email": _clean_text(payload.get("store_email")),
        "market_manager_name": _clean_text(payload.get("market_manager_name")),
        "marketing_phone": _clean_text(payload.get("marketing_phone")),
        "marketing_email": _clean_text(payload.get("marketing_email")),
        "account_dept_name": _clean_text(payload.get("account_dept_name")),
        "accounting_phone": _clean_text(payload.get("accounting_phone")),
        "accounting_email": _clean_text(payload.get("accounting_email")),
        "sort_bucket": _clean_text(payload.get("sort_bucket")),
        "notes": _clean_text(payload.get("notes")),
    }

    if lookup is None:
        db.execute(
            text(
                """
                INSERT INTO stores (
                  customer_id, address_1, address_2, city, state, postcode, country,
                  main_contact, owner_name, owner_phone, owner_email,
                  store_manager_name, store_phone, store_email,
                  market_manager_name, marketing_phone, marketing_email,
                  account_dept_name, accounting_phone, accounting_email,
                  sort_bucket, notes, created_at
                )
                VALUES (
                  :customer_id, NULLIF(:address_1, ''), NULLIF(:address_2, ''), NULLIF(:city, ''), NULLIF(:state, ''),
                  NULLIF(:postcode, ''), NULLIF(:country, ''),
                  NULLIF(:main_contact, ''), NULLIF(:owner_name, ''), NULLIF(:owner_phone, ''), NULLIF(:owner_email, ''),
                  NULLIF(:store_manager_name, ''), NULLIF(:store_phone, ''), NULLIF(:store_email, ''),
                  NULLIF(:market_manager_name, ''), NULLIF(:marketing_phone, ''), NULLIF(:marketing_email, ''),
                  NULLIF(:account_dept_name, ''), NULLIF(:accounting_phone, ''), NULLIF(:accounting_email, ''),
                  NULLIF(:sort_bucket, ''), NULLIF(:notes, ''), NOW()
                )
                """
            ),
            params,
        )
        summary["stores_created"] += 1
        return

    if upsert_policy == "create_only":
        summary["stores_skipped_existing"] += 1
        return

    params["store_id"] = int(lookup["id"])
    if upsert_policy == "overwrite":
        db.execute(
            text(
                """
                UPDATE stores
                SET
                  address_1 = NULLIF(:address_1, ''),
                  address_2 = NULLIF(:address_2, ''),
                  city = NULLIF(:city, ''),
                  state = NULLIF(:state, ''),
                  postcode = NULLIF(:postcode, ''),
                  country = NULLIF(:country, ''),
                  main_contact = NULLIF(:main_contact, ''),
                  owner_name = NULLIF(:owner_name, ''),
                  owner_phone = NULLIF(:owner_phone, ''),
                  owner_email = NULLIF(:owner_email, ''),
                  store_manager_name = NULLIF(:store_manager_name, ''),
                  store_phone = NULLIF(:store_phone, ''),
                  store_email = NULLIF(:store_email, ''),
                  market_manager_name = NULLIF(:market_manager_name, ''),
                  marketing_phone = NULLIF(:marketing_phone, ''),
                  marketing_email = NULLIF(:marketing_email, ''),
                  account_dept_name = NULLIF(:account_dept_name, ''),
                  accounting_phone = NULLIF(:accounting_phone, ''),
                  accounting_email = NULLIF(:accounting_email, ''),
                  sort_bucket = NULLIF(:sort_bucket, ''),
                  notes = NULLIF(:notes, '')
                WHERE id = :store_id
                """
            ),
            params,
        )
        summary["stores_updated"] += 1
        return

    db.execute(
        text(
            """
            UPDATE stores
            SET
              address_1 = COALESCE(NULLIF(:address_1, ''), address_1),
              address_2 = COALESCE(NULLIF(:address_2, ''), address_2),
              city = COALESCE(NULLIF(:city, ''), city),
              state = COALESCE(NULLIF(:state, ''), state),
              postcode = COALESCE(NULLIF(:postcode, ''), postcode),
              country = COALESCE(NULLIF(:country, ''), country),
              main_contact = COALESCE(NULLIF(:main_contact, ''), main_contact),
              owner_name = COALESCE(NULLIF(:owner_name, ''), owner_name),
              owner_phone = COALESCE(NULLIF(:owner_phone, ''), owner_phone),
              owner_email = COALESCE(NULLIF(:owner_email, ''), owner_email),
              store_manager_name = COALESCE(NULLIF(:store_manager_name, ''), store_manager_name),
              store_phone = COALESCE(NULLIF(:store_phone, ''), store_phone),
              store_email = COALESCE(NULLIF(:store_email, ''), store_email),
              market_manager_name = COALESCE(NULLIF(:market_manager_name, ''), market_manager_name),
              marketing_phone = COALESCE(NULLIF(:marketing_phone, ''), marketing_phone),
              marketing_email = COALESCE(NULLIF(:marketing_email, ''), marketing_email),
              account_dept_name = COALESCE(NULLIF(:account_dept_name, ''), account_dept_name),
              accounting_phone = COALESCE(NULLIF(:accounting_phone, ''), accounting_phone),
              accounting_email = COALESCE(NULLIF(:accounting_email, ''), accounting_email),
              sort_bucket = COALESCE(NULLIF(:sort_bucket, ''), sort_bucket),
              notes = COALESCE(NULLIF(:notes, ''), notes)
            WHERE id = :store_id
            """
        ),
        params,
    )
    summary["stores_updated"] += 1


def _upsert_product(
    db: Session,
    summary: dict[str, Any],
    customer_id: int,
    product_name: str,
    upsert_policy: str,
    *,
    last_visit: date | None,
    action: str,
    status: str,
    next_action: str,
    last_contact: date | None,
    notes: str,
) -> None:
    lookup = db.execute(
        text(
            """
            SELECT id
            FROM products
            WHERE customer_id = :customer_id
              AND LOWER(product_name) = LOWER(:product_name)
            ORDER BY id
            LIMIT 1
            """
        ),
        {"customer_id": customer_id, "product_name": product_name},
    ).mappings().first()

    params = {
        "customer_id": customer_id,
        "product_name": product_name,
        "last_visit": last_visit,
        "action": action,
        "status": status,
        "next_action": next_action,
        "last_contact": last_contact,
        "notes": notes,
    }

    if lookup is None:
        db.execute(
            text(
                """
                INSERT INTO products
                  (customer_id, product_name, last_visit, action, status, next_action, last_contact, notes, created_at, updated_at)
                VALUES
                  (:customer_id, :product_name, :last_visit, NULLIF(:action, ''), NULLIF(:status, ''),
                   NULLIF(:next_action, ''), :last_contact, NULLIF(:notes, ''), NOW(), NOW())
                """
            ),
            params,
        )
        summary["products_created"] += 1
        return

    if upsert_policy == "create_only":
        summary["products_skipped_existing"] += 1
        return

    params["product_id"] = int(lookup["id"])
    if upsert_policy == "overwrite":
        db.execute(
            text(
                """
                UPDATE products
                SET
                  last_visit = :last_visit,
                  action = NULLIF(:action, ''),
                  status = NULLIF(:status, ''),
                  next_action = NULLIF(:next_action, ''),
                  last_contact = :last_contact,
                  notes = NULLIF(:notes, ''),
                  updated_at = NOW()
                WHERE id = :product_id
                """
            ),
            params,
        )
        summary["products_updated"] += 1
        return

    db.execute(
        text(
            """
            UPDATE products
            SET
              last_visit = COALESCE(:last_visit, last_visit),
              action = COALESCE(NULLIF(:action, ''), action),
              status = COALESCE(NULLIF(:status, ''), status),
              next_action = COALESCE(NULLIF(:next_action, ''), next_action),
              last_contact = COALESCE(:last_contact, last_contact),
              notes = COALESCE(NULLIF(:notes, ''), notes),
              updated_at = NOW()
            WHERE id = :product_id
            """
        ),
        params,
    )
    summary["products_updated"] += 1


def import_planner_workbook(
    db: Session,
    content: bytes,
    filename: str,
    year_override: int | None = None,
    *,
    upsert_policy: str = "merge",
    validation_mode: str = "standard",
    duplicate_policy: str = "last_wins",
    dry_run: bool = False,
) -> dict[str, Any]:
    if not filename.lower().endswith((".xlsx", ".xlsm", ".xltm")):
        raise HTTPException(status_code=400, detail="Upload an .xlsx or .xlsm workbook")
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded workbook is empty")

    resolved_policy = _normalize_upsert_policy(upsert_policy)
    resolved_validation_mode = _normalize_validation_mode(validation_mode)
    resolved_duplicate_policy = _normalize_duplicate_policy(duplicate_policy)

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail="Workbook import requires openpyxl. Install dependencies and redeploy.",
        ) from exc

    try:
        workbook = load_workbook(BytesIO(content), data_only=True, keep_vba=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read workbook: {exc}") from exc

    summary: dict[str, Any] = {
        "filename": filename,
        "dry_run": bool(dry_run),
        "upsert_policy": resolved_policy,
        "validation_mode": resolved_validation_mode,
        "duplicate_policy": resolved_duplicate_policy,
        "calendar_year": None,
        "territories_created": 0,
        "customers_created": 0,
        "customers_updated": 0,
        "customers_skipped_existing": 0,
        "stores_created": 0,
        "stores_updated": 0,
        "stores_skipped_existing": 0,
        "products_created": 0,
        "products_updated": 0,
        "products_skipped_existing": 0,
        "cvm_entries_upserted": 0,
        "cvm_entries_skipped_existing": 0,
        "warnings": [],
        "warning_count": 0,
        "error_count": 0,
        "row_issues": [],
        "row_issue_limit": ROW_ISSUE_LIMIT,
        "row_issues_truncated": 0,
        "duplicate_rows_skipped": 0,
        "blockers": [],
        "can_apply": True,
    }

    territory_cache: dict[str, int] = {}
    seen_get_data_customer: dict[str, int] = {}
    seen_cvm_customer: dict[str, int] = {}
    seen_database_customer: dict[str, int] = {}
    seen_detail_store: dict[str, int] = {}

    # 1) Get Data sheet -> customers + territory metadata.
    get_data = _sheet_by_prefix(workbook, "Get Data -")
    if get_data is None:
        _add_warning(summary, "Get Data sheet not found; skipped customer master import.")
    else:
        for row_num, row in enumerate(get_data.iter_rows(min_row=2, values_only=True), start=2):
            territory_name = _clean_text(row[0] if len(row) > 0 else None)
            group_name = _clean_text(row[1] if len(row) > 1 else None)
            group_2_iws = _clean_text(row[2] if len(row) > 2 else None)
            iws_code = _clean_text(row[3] if len(row) > 3 else None)
            cust_code = _clean_code(row[4] if len(row) > 4 else None)
            customer_name = _clean_text(row[5] if len(row) > 5 else None)
            old_value = _clean_text(row[6] if len(row) > 6 else None)
            old_name = _clean_text(row[7] if len(row) > 7 else None)

            if not cust_code and not customer_name and not _is_row_populated(row):
                continue
            if not cust_code:
                _record_issue(
                    summary,
                    level="error",
                    sheet=get_data.title,
                    row=row_num,
                    message="Skipped row: missing customer code.",
                )
                continue
            if not customer_name:
                _record_issue(
                    summary,
                    level=_validation_level(summary),
                    sheet=get_data.title,
                    row=row_num,
                    message=f"Customer '{cust_code}' has no customer name; placeholder name may be used.",
                )

            if not _register_duplicate(
                summary,
                seen_get_data_customer,
                key=cust_code,
                sheet=get_data.title,
                row=row_num,
                label="customer",
            ):
                continue

            territory_id = _get_or_create_territory(db, territory_name, territory_cache, summary)
            _upsert_customer(
                db,
                summary,
                cust_code,
                customer_name,
                territory_id,
                resolved_policy,
                group_name=group_name,
                group_2_iws=group_2_iws,
                iws_code=iws_code,
                old_value=old_value,
                old_name=old_name,
            )

    # 2) Customer Details sheet -> stores and richer contact fields.
    details = _sheet_by_prefix(workbook, "Customer Details")
    if details is None:
        _add_warning(summary, "Customer Details sheet not found; skipped store/contact import.")
    else:
        for row_num, row in enumerate(details.iter_rows(min_row=3, values_only=True), start=3):
            cust_code = _clean_code(row[0] if len(row) > 0 else None) or _clean_code(row[2] if len(row) > 2 else None)
            territory_name = _clean_text(row[3] if len(row) > 3 else None)
            customer_name = _extract_name(row[1] if len(row) > 1 else None) or _extract_name(row[2] if len(row) > 2 else None)

            if not cust_code:
                if _is_row_populated(row):
                    _record_issue(
                        summary,
                        level="error",
                        sheet=details.title,
                        row=row_num,
                        message="Skipped row: missing customer code.",
                    )
                continue

            territory_id = _get_or_create_territory(db, territory_name, territory_cache, summary)
            customer_id = _upsert_customer(
                db,
                summary,
                cust_code,
                customer_name,
                territory_id,
                resolved_policy,
            )

            payload = {
                "address_1": row[5] if len(row) > 5 else None,
                "address_2": row[6] if len(row) > 6 else None,
                "city": row[7] if len(row) > 7 else None,
                "state": row[8] if len(row) > 8 else None,
                "postcode": row[9] if len(row) > 9 else None,
                "country": row[10] if len(row) > 10 else None,
                "main_contact": row[11] if len(row) > 11 else None,
                "owner_name": row[12] if len(row) > 12 else None,
                "owner_phone": row[13] if len(row) > 13 else None,
                "owner_email": row[14] if len(row) > 14 else None,
                "store_manager_name": row[15] if len(row) > 15 else None,
                "store_phone": row[16] if len(row) > 16 else None,
                "store_email": row[17] if len(row) > 17 else None,
                "market_manager_name": row[18] if len(row) > 18 else None,
                "marketing_phone": row[19] if len(row) > 19 else None,
                "marketing_email": row[20] if len(row) > 20 else None,
                "account_dept_name": row[21] if len(row) > 21 else None,
                "accounting_phone": row[22] if len(row) > 22 else None,
                "accounting_email": row[23] if len(row) > 23 else None,
                "notes": row[24] if len(row) > 24 else None,
            }
            has_store_data = any(_clean_text(v) for v in payload.values())
            if has_store_data:
                store_key = "|".join(
                    [
                        cust_code.lower(),
                        _clean_text(payload.get("address_1")).lower(),
                        _clean_text(payload.get("city")).lower(),
                        _clean_text(payload.get("state")).lower(),
                    ]
                )
                if not _register_duplicate(
                    summary,
                    seen_detail_store,
                    key=store_key,
                    sheet=details.title,
                    row=row_num,
                    label="store",
                ):
                    continue
                _upsert_store(db, summary, customer_id, payload, resolved_policy)

    # 3) CVM sheet -> monthly planning dates and completion flags.
    cvm = _sheet_by_exact(workbook, "CVM")
    resolved_year = year_override or _resolve_calendar_year(workbook)
    if resolved_year is None:
        resolved_year = datetime.utcnow().year
        _add_warning(
            summary,
            f"Calendar year could not be resolved from workbook. Defaulted to {resolved_year}.",
        )
    calendar_year = resolved_year
    summary["calendar_year"] = int(calendar_year)

    if cvm is None:
        _add_warning(summary, "CVM sheet not found; skipped monthly planning import.")
    else:
        cvm_columns = _resolve_cvm_column_map(cvm)
        for row_num, row in enumerate(cvm.iter_rows(min_row=4, values_only=True), start=4):
            door_count_raw = _column_value(row, cvm_columns["door_count_col"])
            door_count = _to_int(door_count_raw)
            if door_count is None and _clean_text(door_count_raw):
                _record_issue(
                    summary,
                    level=_validation_level(summary),
                    sheet=cvm.title,
                    row=row_num,
                    message=f"Invalid Door Count '{_clean_text(door_count_raw)}'; value ignored.",
                )

            cust_code = _clean_code(_column_value(row, cvm_columns["cust_code_col"]))
            territory_name = _clean_text(_column_value(row, cvm_columns["territory_col"]))
            customer_name = _clean_text(_column_value(row, cvm_columns["customer_name_col"]))
            trade_name = _clean_text(_column_value(row, cvm_columns["trade_name_col"]))
            cvm_notes = _clean_text(_column_value(row, cvm_columns["notes_col"]))
            sort_bucket = _clean_text(_column_value(row, cvm_columns["sort_col"]))

            row_signal_values = [
                _column_value(row, cvm_columns["territory_col"]),
                _column_value(row, cvm_columns["customer_name_col"]),
                _column_value(row, cvm_columns["trade_name_col"]),
                _column_value(row, cvm_columns["notes_col"]),
            ]
            for month_idx in range(1, 13):
                month_columns = cvm_columns["month_cols"].get(month_idx, {})
                row_signal_values.append(_column_value(row, month_columns.get("planned_col")))
                row_signal_values.append(_column_value(row, month_columns.get("completed_col")))

            if not cust_code:
                if _has_meaningful_values(row_signal_values):
                    _record_issue(
                        summary,
                        level="error",
                        sheet=cvm.title,
                        row=row_num,
                        message="Skipped row: missing customer code.",
                    )
                continue

            if not _register_duplicate(
                summary,
                seen_cvm_customer,
                key=cust_code,
                sheet=cvm.title,
                row=row_num,
                label="customer",
            ):
                continue

            territory_id = _get_or_create_territory(db, territory_name, territory_cache, summary)
            customer_id = _upsert_customer(
                db,
                summary,
                cust_code,
                customer_name,
                territory_id,
                resolved_policy,
                trade_name=trade_name,
                door_count=door_count,
                cvm_notes=cvm_notes,
            )

            if sort_bucket:
                db.execute(
                    text(
                        """
                        UPDATE stores
                        SET sort_bucket = COALESCE(NULLIF(:sort_bucket, ''), sort_bucket)
                        WHERE id = (
                          SELECT id
                          FROM stores
                          WHERE customer_id = :customer_id
                          ORDER BY id
                          LIMIT 1
                        )
                        """
                    ),
                    {"customer_id": customer_id, "sort_bucket": sort_bucket},
                )

            for month_idx in range(1, 13):
                month_columns = cvm_columns["month_cols"].get(month_idx, {})
                planned_col = month_columns.get("planned_col")
                completed_col = month_columns.get("completed_col")
                planned_date = _to_date_with_issue(
                    _column_value(row, planned_col),
                    summary=summary,
                    sheet=cvm.title,
                    row=row_num,
                    field=f"PLANNED {MONTH_SHORT[month_idx - 1]}",
                )
                completed_manual = _to_bool(_column_value(row, completed_col))
                if completed_manual and not planned_date:
                    _record_issue(
                        summary,
                        level=_validation_level(summary),
                        sheet=cvm.title,
                        row=row_num,
                        message=f"COMPLETED {MONTH_SHORT[month_idx - 1]} ignored because planned date is missing or invalid.",
                    )
                    completed_manual = False
                if not planned_date and not completed_manual:
                    continue

                params = {
                    "customer_id": customer_id,
                    "year": int(calendar_year),
                    "month": month_idx,
                    "planned_date": planned_date,
                    "completed_manual": completed_manual,
                }
                if resolved_policy == "create_only":
                    exists = db.execute(
                        text(
                            """
                            SELECT 1
                            FROM cvm_month_entries
                            WHERE customer_id = :customer_id
                              AND year = :year
                              AND month = :month
                            """
                        ),
                        params,
                    ).scalar()
                    if exists:
                        summary["cvm_entries_skipped_existing"] += 1
                        continue
                    db.execute(
                        text(
                            """
                            INSERT INTO cvm_month_entries
                              (customer_id, year, month, planned_date, completed_manual, updated_at)
                            VALUES
                              (:customer_id, :year, :month, :planned_date, :completed_manual, NOW())
                            """
                        ),
                        params,
                    )
                    summary["cvm_entries_upserted"] += 1
                    continue

                db.execute(
                    text(
                        """
                        INSERT INTO cvm_month_entries
                          (customer_id, year, month, planned_date, completed_manual, updated_at)
                        VALUES
                          (:customer_id, :year, :month, :planned_date, :completed_manual, NOW())
                        ON CONFLICT (customer_id, year, month)
                        DO UPDATE SET
                          planned_date = EXCLUDED.planned_date,
                          completed_manual = EXCLUDED.completed_manual,
                          updated_at = NOW()
                        """
                    ),
                    params,
                )
                summary["cvm_entries_upserted"] += 1

    # 4) Database sheet -> product interaction snapshot(s).
    database = _sheet_by_exact(workbook, "Database")
    if database is None:
        _add_warning(summary, "Database sheet not found; skipped product import.")
    else:
        max_col = database.max_column
        product_labels = [_clean_text(database.cell(row=3, column=i).value) for i in range(1, max_col + 1)]
        field_labels = [_clean_text(database.cell(row=4, column=i).value) for i in range(1, max_col + 1)]
        database_columns = _resolve_database_column_map(database)

        product_groups: list[tuple[int, str]] = []
        for i, field_name in enumerate(field_labels, start=1):
            if field_name.upper().startswith("ACTION") and i + 4 <= max_col:
                product_name = product_labels[i - 1] or f"Product {len(product_groups) + 1}"
                product_groups.append((i, product_name))

        if not product_groups:
            _add_warning(summary, "No ACTION product groups found in Database sheet.")
        else:
            for row_idx in range(5, database.max_row + 1):
                cust_code_raw = _cell_value(database, row_idx, database_columns["cust_code_col"])
                territory_raw = _cell_value(database, row_idx, database_columns["territory_col"])
                customer_name_raw = _cell_value(database, row_idx, database_columns["customer_name_col"])
                trade_name_raw = _cell_value(database, row_idx, database_columns["trade_name_col"])
                last_visit_raw = _cell_value(database, row_idx, database_columns["last_visit_col"])

                cust_code = _clean_code(cust_code_raw)
                territory_name = _clean_text(territory_raw)
                customer_name = _clean_text(customer_name_raw)
                trade_name = _clean_text(trade_name_raw)
                last_visit = _to_date_with_issue(
                    last_visit_raw,
                    summary=summary,
                    sheet=database.title,
                    row=row_idx,
                    field="LAST VISIT",
                )

                row_signal_values = [territory_raw, customer_name_raw, trade_name_raw, last_visit_raw]
                for action_col, _product_name in product_groups:
                    row_signal_values.extend(
                        [
                            _cell_value(database, row_idx, action_col),
                            _cell_value(database, row_idx, action_col + 1),
                            _cell_value(database, row_idx, action_col + 2),
                            _cell_value(database, row_idx, action_col + 3),
                            _cell_value(database, row_idx, action_col + 4),
                        ]
                    )

                if not cust_code:
                    if _has_meaningful_values(row_signal_values):
                        _record_issue(
                            summary,
                            level="error",
                            sheet=database.title,
                            row=row_idx,
                            message="Skipped row: missing customer code.",
                        )
                    continue

                if not _register_duplicate(
                    summary,
                    seen_database_customer,
                    key=cust_code,
                    sheet=database.title,
                    row=row_idx,
                    label="customer",
                ):
                    continue

                territory_id = _get_or_create_territory(db, territory_name, territory_cache, summary)
                customer_id = _upsert_customer(
                    db,
                    summary,
                    cust_code,
                    customer_name,
                    territory_id,
                    resolved_policy,
                    trade_name=trade_name,
                )

                for action_col, product_name in product_groups:
                    action = _clean_text(database.cell(row=row_idx, column=action_col).value)
                    status = _clean_text(database.cell(row=row_idx, column=action_col + 1).value)
                    next_action = _clean_text(database.cell(row=row_idx, column=action_col + 2).value)
                    last_contact = _to_date_with_issue(
                        database.cell(row=row_idx, column=action_col + 3).value,
                        summary=summary,
                        sheet=database.title,
                        row=row_idx,
                        field=f"{product_name} LAST CONTACT",
                    )
                    notes = _clean_text(database.cell(row=row_idx, column=action_col + 4).value)

                    if not any([action, status, next_action, last_contact, notes, last_visit]):
                        continue

                    _upsert_product(
                        db,
                        summary,
                        customer_id,
                        product_name=product_name,
                        upsert_policy=resolved_policy,
                        last_visit=last_visit,
                        action=action,
                        status=status,
                        next_action=next_action,
                        last_contact=last_contact,
                        notes=notes,
                    )

    if summary["validation_mode"] == "strict" and int(summary.get("error_count", 0)) > 0:
        _add_blocker(
            summary,
            f"Strict validation found {summary['error_count']} error(s). Resolve errors before applying import.",
        )

    summary["can_apply"] = len(summary.get("blockers", [])) == 0
    return summary
