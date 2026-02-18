from datetime import datetime
from contextlib import asynccontextmanager
from pathlib import Path
from io import BytesIO
import sys

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, event, text
from sqlalchemy.orm import sessionmaker

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app import main


@pytest.fixture()
def client_and_engine(tmp_path):
    db_path = tmp_path / "test.db"
    engine = create_engine(
        f"sqlite:///{db_path}",
        connect_args={"check_same_thread": False},
    )

    @event.listens_for(engine, "connect")
    def register_now(dbapi_conn, _):
        dbapi_conn.create_function("NOW", 0, lambda: datetime.utcnow().isoformat(sep=" "))

    SessionTesting = sessionmaker(bind=engine, autoflush=False, autocommit=False)

    with engine.begin() as conn:
        conn.execute(text("CREATE TABLE territories (id INTEGER PRIMARY KEY, name TEXT UNIQUE)"))
        conn.execute(
            text(
                """
                CREATE TABLE customers (
                  id INTEGER PRIMARY KEY,
                  cust_code TEXT UNIQUE,
                  name TEXT NOT NULL,
                  trade_name TEXT,
                  territory_id INTEGER,
                  group_name TEXT,
                  group_2_iws TEXT,
                  iws_code TEXT,
                  old_value TEXT,
                  old_name TEXT,
                  door_count INTEGER,
                  cvm_notes TEXT,
                  created_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE stores (
                  id INTEGER PRIMARY KEY,
                  customer_id INTEGER,
                  address_1 TEXT,
                  address_2 TEXT,
                  city TEXT,
                  state TEXT,
                  postcode TEXT,
                  country TEXT,
                  main_contact TEXT,
                  owner_name TEXT,
                  owner_phone TEXT,
                  owner_email TEXT,
                  store_manager_name TEXT,
                  store_phone TEXT,
                  store_email TEXT,
                  market_manager_name TEXT,
                  marketing_phone TEXT,
                  marketing_email TEXT,
                  account_dept_name TEXT,
                  accounting_phone TEXT,
                  accounting_email TEXT,
                  sort_bucket TEXT,
                  notes TEXT,
                  created_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE visit_events (
                  id INTEGER PRIMARY KEY,
                  customer_id INTEGER,
                  store_id INTEGER,
                  event_type TEXT,
                  event_date DATE,
                  action TEXT,
                  status TEXT,
                  next_action TEXT,
                  last_contact DATE,
                  notes TEXT,
                  created_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE cvm_month_entries (
                  id INTEGER PRIMARY KEY,
                  customer_id INTEGER NOT NULL,
                  year INTEGER NOT NULL,
                  month INTEGER NOT NULL,
                  planned_date DATE,
                  completed_manual BOOLEAN NOT NULL DEFAULT 0,
                  updated_at TEXT,
                  UNIQUE(customer_id, year, month)
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE products (
                  id INTEGER PRIMARY KEY,
                  customer_id INTEGER NOT NULL,
                  product_name TEXT NOT NULL,
                  last_visit DATE,
                  action TEXT,
                  status TEXT,
                  next_action TEXT,
                  last_contact DATE,
                  notes TEXT,
                  created_at TEXT,
                  updated_at TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE reference_values (
                  id INTEGER PRIMARY KEY,
                  category TEXT NOT NULL,
                  value TEXT NOT NULL,
                  sort_order INTEGER NOT NULL DEFAULT 0,
                  active BOOLEAN NOT NULL DEFAULT 1
                )
                """
            )
        )

    def override_get_db():
        db = SessionTesting()
        try:
            yield db
        finally:
            db.close()

    original_startup = list(main.app.router.on_startup)
    original_shutdown = list(main.app.router.on_shutdown)
    original_lifespan = main.app.router.lifespan_context

    
    async def _noop_lifespan(_app):
        yield

    main.app.router.on_startup = []
    main.app.router.on_shutdown = []
    main.app.router.lifespan_context = _noop_lifespan
    main.app.dependency_overrides[main.get_db] = override_get_db

    client = TestClient(main.app)
    try:
        yield client, engine
    finally:
        client.close()
        main.app.dependency_overrides.clear()
        main.app.router.on_startup = original_startup
        main.app.router.on_shutdown = original_shutdown
        main.app.router.lifespan_context = original_lifespan
        engine.dispose()


def seed_customer(engine, customer_id: int, code: str, name: str):
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO customers (id, cust_code, name, trade_name, created_at)
                VALUES (:id, :code, :name, :trade_name, NOW())
                """
            ),
            {"id": customer_id, "code": code, "name": name, "trade_name": "Trade"},
        )


def test_create_product_rejects_unknown_customer(client_and_engine):
    client, _ = client_and_engine

    response = client.post(
        "/products",
        data={
            "customer_id": "999",
            "product_name": "Test Product",
            "last_visit": "",
            "action": "",
            "status": "",
            "next_action": "",
            "last_contact": "",
            "notes": "",
        },
        follow_redirects=False,
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Invalid customer_id"


def test_create_product_rejects_invalid_last_visit(client_and_engine):
    client, engine = client_and_engine
    seed_customer(engine, 1, "C1", "Customer 1")

    response = client.post(
        "/products",
        data={
            "customer_id": "1",
            "product_name": "Test Product",
            "last_visit": "bad-date",
            "action": "",
            "status": "",
            "next_action": "",
            "last_contact": "",
            "notes": "",
        },
        follow_redirects=False,
    )

    assert response.status_code == 400
    assert "Invalid last_visit" in response.json()["detail"]


def test_update_product_rejects_unknown_product(client_and_engine):
    client, engine = client_and_engine
    seed_customer(engine, 1, "C1", "Customer 1")

    response = client.post(
        "/products/999",
        data={
            "customer_id": "1",
            "product_name": "Test Product",
            "last_visit": "",
            "action": "",
            "status": "",
            "next_action": "",
            "last_contact": "",
            "notes": "",
        },
        follow_redirects=False,
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Invalid product_id"


def test_create_product_rejects_invalid_last_contact(client_and_engine):
    client, engine = client_and_engine
    seed_customer(engine, 1, "C1", "Customer 1")

    response = client.post(
        "/products",
        data={
            "customer_id": "1",
            "product_name": "Test Product",
            "last_visit": "",
            "action": "",
            "status": "",
            "next_action": "",
            "last_contact": "not-a-date",
            "notes": "",
        },
        follow_redirects=False,
    )

    assert response.status_code == 400
    assert "Invalid last_contact" in response.json()["detail"]


def test_create_product_success_persists(client_and_engine):
    client, engine = client_and_engine
    seed_customer(engine, 1, "C1", "Customer 1")

    response = client.post(
        "/products",
        data={
            "customer_id": "1",
            "product_name": "Test Product",
            "last_visit": "2026-02-01",
            "action": "Call",
            "status": "Open",
            "next_action": "Follow Up",
            "last_contact": "2026-02-02",
            "notes": "Note one",
        },
        follow_redirects=False,
    )

    assert response.status_code == 303
    assert response.headers["location"] == "/products"

    with engine.begin() as conn:
        row = conn.execute(
            text(
                """
                SELECT customer_id, product_name, last_visit, action, status, next_action, last_contact, notes
                FROM products
                WHERE customer_id = 1 AND product_name = 'Test Product'
                """
            )
        ).mappings().first()

    assert row is not None
    assert row["customer_id"] == 1
    assert row["action"] == "Call"
    assert row["status"] == "Open"
    assert row["next_action"] == "Follow Up"
    assert str(row["last_visit"]).startswith("2026-02-01")
    assert str(row["last_contact"]).startswith("2026-02-02")
    assert row["notes"] == "Note one"


def test_update_product_rejects_invalid_last_contact(client_and_engine):
    client, engine = client_and_engine
    seed_customer(engine, 1, "C1", "Customer 1")
    seed_product(engine, 1, 1, "Existing Product")

    response = client.post(
        "/products/1",
        data={
            "customer_id": "1",
            "product_name": "Existing Product",
            "last_visit": "",
            "action": "",
            "status": "",
            "next_action": "",
            "last_contact": "bad-contact-date",
            "notes": "",
        },
        follow_redirects=False,
    )

    assert response.status_code == 400
    assert "Invalid last_contact" in response.json()["detail"]


def test_update_product_success_persists(client_and_engine):
    client, engine = client_and_engine
    seed_customer(engine, 1, "C1", "Customer 1")
    seed_product(engine, 1, 1, "Existing Product")

    response = client.post(
        "/products/1",
        data={
            "customer_id": "1",
            "product_name": "Updated Product",
            "last_visit": "2026-03-01",
            "action": "Visit",
            "status": "Closed",
            "next_action": "Review",
            "last_contact": "2026-03-03",
            "notes": "Updated note",
        },
        follow_redirects=False,
    )

    assert response.status_code == 303
    assert response.headers["location"] == "/products"

    with engine.begin() as conn:
        row = conn.execute(
            text(
                """
                SELECT product_name, last_visit, action, status, next_action, last_contact, notes
                FROM products
                WHERE id = 1
                """
            )
        ).mappings().first()

    assert row is not None
    assert row["product_name"] == "Updated Product"
    assert row["action"] == "Visit"
    assert row["status"] == "Closed"
    assert row["next_action"] == "Review"
    assert str(row["last_visit"]).startswith("2026-03-01")
    assert str(row["last_contact"]).startswith("2026-03-03")
    assert row["notes"] == "Updated note"


def test_create_customer_duplicate_renders_form_error(client_and_engine):
    client, engine = client_and_engine
    seed_customer(engine, 1, "C1", "Customer One")

    response = client.post(
        "/customers",
        data={
            "cust_code": "C1",
            "name": "Customer One Duplicate",
            "trade_name": "Dup Trade",
            "territory_name": "NSW",
            "group_name": "Group A",
            "iws_code": "IWS100",
        },
        follow_redirects=False,
    )

    assert response.status_code == 400
    assert "Could not create customer. Check for duplicate customer code or invalid values." in response.text
    assert 'name="cust_code"' in response.text
    assert 'value="C1"' in response.text
    assert 'value="Customer One Duplicate"' in response.text


def test_create_customer_blank_required_fields_renders_form_error(client_and_engine):
    client, _ = client_and_engine

    response = client.post(
        "/customers",
        data={
            "cust_code": "   ",
            "name": "   ",
            "trade_name": "",
            "territory_name": "",
            "group_name": "",
            "iws_code": "",
        },
        follow_redirects=False,
    )

    assert response.status_code == 400
    assert "Cust Code and Customer Name are required." in response.text


def test_delete_customer_removes_customer_and_related_data(client_and_engine):
    client, engine = client_and_engine
    seed_customer(engine, 1, "C1", "Customer 1")

    with engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO stores (id, customer_id, address_1, created_at)
                VALUES (1, 1, '1 Test St', NOW())
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO products (id, customer_id, product_name, created_at, updated_at)
                VALUES (1, 1, 'Test Product', NOW(), NOW())
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO cvm_month_entries (id, customer_id, year, month, planned_date, completed_manual, updated_at)
                VALUES (1, 1, 2026, 2, '2026-02-10', 1, NOW())
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO visit_events (id, customer_id, store_id, event_type, event_date, created_at)
                VALUES (1, 1, 1, 'planned', '2026-02-10', NOW())
                """
            )
        )

    response = client.post("/customers/1/delete", follow_redirects=False)

    assert response.status_code == 303
    assert response.headers["location"] == "/customers"

    with engine.begin() as conn:
        customer_count = conn.execute(text("SELECT COUNT(*) FROM customers WHERE id = 1")).scalar_one()
        stores_count = conn.execute(text("SELECT COUNT(*) FROM stores WHERE customer_id = 1")).scalar_one()
        products_count = conn.execute(text("SELECT COUNT(*) FROM products WHERE customer_id = 1")).scalar_one()
        cvm_count = conn.execute(text("SELECT COUNT(*) FROM cvm_month_entries WHERE customer_id = 1")).scalar_one()
        events_count = conn.execute(text("SELECT COUNT(*) FROM visit_events WHERE customer_id = 1")).scalar_one()

    assert customer_count == 0
    assert stores_count == 0
    assert products_count == 0
    assert cvm_count == 0
    assert events_count == 0


def test_delete_customer_rejects_unknown_customer(client_and_engine):
    client, _ = client_and_engine

    response = client.post("/customers/999/delete", follow_redirects=False)

    assert response.status_code == 400
    assert response.json()["detail"] == "Invalid customer_id"


def test_cvm_month_update_rejects_unknown_customer(client_and_engine):
    client, _ = client_and_engine

    response = client.post(
        "/cvm/month-update",
        data={
            "customer_id": "999",
            "year": "2026",
            "month": "2",
            "planned_date": "2026-02-12",
            "completed_manual": "1",
            "territory_id": "",
        },
        follow_redirects=False,
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Invalid customer_id"


def test_cvm_month_update_rejects_invalid_date(client_and_engine):
    client, engine = client_and_engine
    seed_customer(engine, 1, "C1", "Customer 1")

    response = client.post(
        "/cvm/month-update",
        data={
            "customer_id": "1",
            "year": "2026",
            "month": "2",
            "planned_date": "2026-99-12",
            "completed_manual": "1",
            "territory_id": "",
        },
        follow_redirects=False,
    )

    assert response.status_code == 400
    assert "Invalid planned_date" in response.json()["detail"]


def test_cvm_month_update_success_persists_manual_done(client_and_engine):
    client, engine = client_and_engine
    seed_customer(engine, 1, "C1", "Customer 1")

    response = client.post(
        "/cvm/month-update",
        data={
            "customer_id": "1",
            "year": "2026",
            "month": "8",
            "planned_date": "2026-08-14",
            "completed_manual": "1",
            "territory_id": "",
        },
        follow_redirects=False,
    )

    assert response.status_code == 303

    with engine.begin() as conn:
        row = conn.execute(
            text(
                """
                SELECT planned_date, completed_manual
                FROM cvm_month_entries
                WHERE customer_id = 1 AND year = 2026 AND month = 8
                """
            )
        ).mappings().first()

    assert row is not None
    assert str(row["planned_date"]).startswith("2026-08-14")
    assert bool(row["completed_manual"]) is True


def seed_territory(engine, territory_id: int, name: str):
    with engine.begin() as conn:
        conn.execute(
            text("INSERT INTO territories (id, name) VALUES (:id, :name)"),
            {"id": territory_id, "name": name},
        )


def seed_reference_value(engine, category: str, value: str, sort_order: int = 1):
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO reference_values (category, value, sort_order, active)
                VALUES (:category, :value, :sort_order, 1)
                """
            ),
            {"category": category, "value": value, "sort_order": sort_order},
        )


def seed_product(
    engine,
    product_id: int,
    customer_id: int,
    product_name: str,
    action: str | None = None,
    status: str | None = None,
    notes: str | None = None,
):
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO products
                  (id, customer_id, product_name, action, status, notes, created_at, updated_at)
                VALUES
                  (:id, :customer_id, :product_name, :action, :status, :notes, NOW(), NOW())
                """
            ),
            {
                "id": product_id,
                "customer_id": customer_id,
                "product_name": product_name,
                "action": action,
                "status": status,
                "notes": notes,
            },
        )


def test_products_add_form_orders_product_before_customer(client_and_engine):
    client, engine = client_and_engine
    seed_customer(engine, 1, "C1", "Customer One")

    response = client.get("/products")

    assert response.status_code == 200
    page = response.text
    form_start = page.find('<form method="post" action="/products" class="grid">')
    form_end = page.find("</form>", form_start)
    assert form_start >= 0 and form_end > form_start

    add_form = page[form_start:form_end]
    assert add_form.find('name="product_name"') < add_form.find('name="customer_id"')


def test_products_page_filters_results(client_and_engine):
    client, engine = client_and_engine

    seed_territory(engine, 1, "North")
    seed_territory(engine, 2, "South")
    seed_customer(engine, 1, "C1", "Customer One")
    seed_customer(engine, 2, "C2", "Customer Two")

    with engine.begin() as conn:
        conn.execute(text("UPDATE customers SET territory_id = 1 WHERE id = 1"))
        conn.execute(text("UPDATE customers SET territory_id = 2 WHERE id = 2"))

    seed_reference_value(engine, "action", "Call", 1)
    seed_reference_value(engine, "action", "Visit", 2)
    seed_reference_value(engine, "status", "Open", 1)
    seed_reference_value(engine, "status", "Closed", 2)

    seed_product(engine, 1, 1, "Widget Alpha", action="Call", status="Open", notes="priority")
    seed_product(engine, 2, 1, "Widget Beta", action="Visit", status="Open")
    seed_product(engine, 3, 2, "Gadget Zeta", action="Call", status="Closed")

    response = client.get(
        "/products",
        params={
            "customer_id": "1",
            "territory": "North",
            "action": "Call",
            "status": "Open",
            "q": "alpha",
        },
    )

    assert response.status_code == 200
    body = response.text
    assert "Widget Alpha" in body
    assert "Widget Beta" not in body
    assert "Gadget Zeta" not in body


def _build_minimal_workbook_bytes() -> bytes:
    if Workbook is None:
        pytest.skip("openpyxl is not installed in this test environment")

    wb = Workbook()

    get_data = wb.active
    get_data.title = "Get Data -Sample"
    get_data.append(["Territory", "Group", "Group 2 IWS", "IWS Codes", "Customer Number", "Customer Name", "OLD Value", "Old Name"])
    get_data.append(["NSW (North)", "IWS", "WORKLOCKER", "IWS001", "C100", "Alpha Store", "", ""])

    customer_details = wb.create_sheet("Customer Details ")
    customer_details.append([])
    customer_details.append(["Custom", "Customer Object", "Customer Object (BillTo)", "Customer Territory A Name", "Customer Territory B Name", "STORE ADDRESS 1"])
    customer_details.append(
        [
            "C100",
            "Alpha Store",
            "C100",
            "NSW (North)",
            "",
            "1 Test St",
            "",
            "Newcastle",
            "NSW",
            "2300",
            "AUSTRALIA",
            "Main Contact",
            "Owner Name",
            "0400000000",
            "owner@example.com",
            "Store Manager",
            "0200000000",
            "store@example.com",
            "Market Manager",
            "0300000000",
            "marketing@example.com",
            "Accounts",
            "0400000001",
            "accounts@example.com",
            "Store notes",
        ]
    )

    cvm = wb.create_sheet(" CVM")
    cvm.append([])
    cvm.append([])
    cvm.append(
        [
            "Door Count",
            "TERRITORY",
            "Cust Code",
            "SORT",
            "Customer Name",
            "TRADE NAME",
            "Notes/Comments",
            "Date of last completed visit",
            "Count of Planned",
            "Count of Visit",
            "PLANNED JAN",
            "COMPLETED JAN",
        ]
    )
    cvm.append(
        [
            1,
            "NSW (North)",
            "C100",
            "northc",
            "Alpha Store",
            "Alpha Trade",
            "Needs monthly visit",
            datetime(2026, 1, 10),
            1,
            1,
            datetime(2026, 1, 20),
            True,
        ]
    )

    january = wb.create_sheet("JANUARY")
    january["R4"] = 2026

    database = wb.create_sheet("Database")
    database.cell(row=3, column=28).value = "DURA X SELL IN"
    for offset, label in enumerate(["ACTION", "STATUS", "NEXT ACTION", "LAST CONTACT", "NOTES"]):
        database.cell(row=4, column=28 + offset).value = label

    database.cell(row=5, column=21).value = "NSW (North)"
    database.cell(row=5, column=22).value = "C100"
    database.cell(row=5, column=23).value = "Alpha Store"
    database.cell(row=5, column=24).value = "Alpha Trade"
    database.cell(row=5, column=25).value = datetime(2026, 1, 5)
    database.cell(row=5, column=28).value = "CALL"
    database.cell(row=5, column=29).value = "ORDERED"
    database.cell(row=5, column=30).value = "Follow up"
    database.cell(row=5, column=31).value = datetime(2026, 1, 6)
    database.cell(row=5, column=32).value = "Imported note"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_invalid_date_workbook_bytes() -> bytes:
    if Workbook is None:
        pytest.skip("openpyxl is not installed in this test environment")

    wb = Workbook()
    get_data = wb.active
    get_data.title = "Get Data -Sample"
    get_data.append(["Territory", "Group", "Group 2 IWS", "IWS Codes", "Customer Number", "Customer Name", "OLD Value", "Old Name"])
    get_data.append(["NSW (North)", "IWS", "", "", "C101", "Invalid Date Store", "", ""])

    cvm = wb.create_sheet("CVM")
    cvm.append([])
    cvm.append([])
    cvm.append(
        [
            "Door Count",
            "TERRITORY",
            "Cust Code",
            "SORT",
            "Customer Name",
            "TRADE NAME",
            "Notes/Comments",
            "Date of last completed visit",
            "Count of Planned",
            "Count of Visit",
            "PLANNED JAN",
            "COMPLETED JAN",
        ]
    )
    cvm.append([1, "NSW (North)", "C101", "northc", "Invalid Date Store", "Trade", "note", "", "", "", "bad-date", True])

    january = wb.create_sheet("JANUARY")
    january["R4"] = 2026

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_duplicate_customer_workbook_bytes() -> bytes:
    if Workbook is None or load_workbook is None:
        pytest.skip("openpyxl is not installed in this test environment")

    wb = load_workbook(BytesIO(_build_minimal_workbook_bytes()), keep_vba=True)
    get_data = wb["Get Data -Sample"]
    get_data.append(["NSW (North)", "IWS", "WORKLOCKER", "IWS001", "C100", "Alpha Store Duplicate", "", ""])

    cvm = wb["CVM"] if "CVM" in wb.sheetnames else wb[" CVM"]
    cvm.append(
        [
            1,
            "NSW (North)",
            "C100",
            "northc",
            "Alpha Store Duplicate",
            "Alpha Trade",
            "Duplicate row",
            datetime(2026, 1, 11),
            1,
            1,
            datetime(2026, 1, 21),
            True,
        ]
    )

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def test_cvm_notes_update_persists(client_and_engine):
    client, engine = client_and_engine
    seed_customer(engine, 1, "C1", "Customer 1")

    response = client.post(
        "/cvm/notes-update",
        data={
            "customer_id": "1",
            "notes": "Call quarterly",
            "year": "2026",
            "territory_id": "",
        },
        follow_redirects=False,
    )

    assert response.status_code == 303

    with engine.begin() as conn:
        note = conn.execute(text("SELECT cvm_notes FROM customers WHERE id = 1")).scalar_one()

    assert note == "Call quarterly"


def test_import_workbook_ingests_core_data(client_and_engine):
    client, engine = client_and_engine

    workbook_bytes = _build_minimal_workbook_bytes()

    response = client.post(
        "/import/workbook",
        data={"year_override": "2026"},
        files={
            "workbook_file": (
                "planner.xlsm",
                workbook_bytes,
                "application/vnd.ms-excel.sheet.macroEnabled.12",
            )
        },
    )

    assert response.status_code == 200
    assert "Import Summary" in response.text

    with engine.begin() as conn:
        customer = conn.execute(
            text("SELECT id, cvm_notes FROM customers WHERE cust_code = 'C100'")
        ).mappings().first()
        assert customer is not None
        assert customer["cvm_notes"] == "Needs monthly visit"

        store_count = conn.execute(
            text("SELECT COUNT(*) FROM stores WHERE customer_id = :customer_id"),
            {"customer_id": customer["id"]},
        ).scalar_one()
        assert store_count >= 1

        cvm_entry = conn.execute(
            text(
                """
                SELECT planned_date, completed_manual
                FROM cvm_month_entries
                WHERE customer_id = :customer_id AND year = 2026 AND month = 1
                """
            ),
            {"customer_id": customer["id"]},
        ).mappings().first()
        assert cvm_entry is not None
        assert str(cvm_entry["planned_date"]).startswith("2026-01-20")
        assert bool(cvm_entry["completed_manual"]) is True

        product = conn.execute(
            text(
                """
                SELECT product_name, action, status
                FROM products
                WHERE customer_id = :customer_id
                """
            ),
            {"customer_id": customer["id"]},
        ).mappings().first()
        assert product is not None
        assert product["product_name"] == "DURA X SELL IN"
        assert product["action"] == "CALL"
        assert product["status"] == "ORDERED"


def test_import_workbook_preview_does_not_commit(client_and_engine):
    client, engine = client_and_engine
    workbook_bytes = _build_minimal_workbook_bytes()

    response = client.post(
        "/import/workbook",
        data={
            "year_override": "2026",
            "import_mode": "preview",
            "upsert_policy": "merge",
        },
        files={
            "workbook_file": (
                "planner.xlsm",
                workbook_bytes,
                "application/vnd.ms-excel.sheet.macroEnabled.12",
            )
        },
    )

    assert response.status_code == 200
    assert "Preview only" in response.text

    with engine.begin() as conn:
        customer_count = conn.execute(
            text("SELECT COUNT(*) FROM customers WHERE cust_code = 'C100'")
        ).scalar_one()
    assert customer_count == 0


def test_import_workbook_create_only_skips_existing_customer_updates(client_and_engine):
    client, engine = client_and_engine
    workbook_bytes = _build_minimal_workbook_bytes()

    seed_customer(engine, 1, "C100", "Original Name")
    with engine.begin() as conn:
        conn.execute(
            text("UPDATE customers SET cvm_notes = 'Keep this note' WHERE id = 1")
        )

    response = client.post(
        "/import/workbook",
        data={
            "year_override": "2026",
            "import_mode": "apply",
            "upsert_policy": "create_only",
        },
        files={
            "workbook_file": (
                "planner.xlsm",
                workbook_bytes,
                "application/vnd.ms-excel.sheet.macroEnabled.12",
            )
        },
    )

    assert response.status_code == 200
    assert "Customers Skipped Existing" in response.text

    with engine.begin() as conn:
        row = conn.execute(
            text("SELECT name, cvm_notes FROM customers WHERE cust_code = 'C100'")
        ).mappings().first()

    assert row is not None
    assert row["name"] == "Original Name"
    assert row["cvm_notes"] == "Keep this note"


def test_import_workbook_apply_strict_blocks_on_row_errors(client_and_engine):
    client, engine = client_and_engine
    workbook_bytes = _build_invalid_date_workbook_bytes()

    response = client.post(
        "/import/workbook",
        data={
            "year_override": "2026",
            "import_mode": "apply",
            "upsert_policy": "merge",
            "validation_mode": "strict",
            "duplicate_policy": "last_wins",
        },
        files={
            "workbook_file": (
                "planner-invalid.xlsm",
                workbook_bytes,
                "application/vnd.ms-excel.sheet.macroEnabled.12",
            )
        },
    )

    assert response.status_code == 400
    assert "Import blocked by validation rules" in response.text
    assert "Apply Blockers" in response.text

    with engine.begin() as conn:
        customer_count = conn.execute(
            text("SELECT COUNT(*) FROM customers WHERE cust_code = 'C101'")
        ).scalar_one()
    assert customer_count == 0


def test_import_workbook_ignores_completed_without_valid_planned_date(client_and_engine):
    client, engine = client_and_engine
    workbook_bytes = _build_invalid_date_workbook_bytes()

    response = client.post(
        "/import/workbook",
        data={
            "year_override": "2026",
            "import_mode": "apply",
            "upsert_policy": "merge",
            "validation_mode": "standard",
            "duplicate_policy": "last_wins",
        },
        files={
            "workbook_file": (
                "planner-invalid.xlsm",
                workbook_bytes,
                "application/vnd.ms-excel.sheet.macroEnabled.12",
            )
        },
    )

    assert response.status_code == 200
    assert "COMPLETED JAN ignored because planned date is missing or invalid." in response.text

    with engine.begin() as conn:
        customer = conn.execute(
            text("SELECT id FROM customers WHERE cust_code = 'C101'")
        ).mappings().first()
        assert customer is not None
        cvm_count = conn.execute(
            text(
                """
                SELECT COUNT(*)
                FROM cvm_month_entries
                WHERE customer_id = :customer_id
                  AND year = 2026
                  AND month = 1
                """
            ),
            {"customer_id": customer["id"]},
        ).scalar_one()

    assert cvm_count == 0


def test_import_workbook_duplicate_policy_error_blocks_apply(client_and_engine):
    client, engine = client_and_engine
    workbook_bytes = _build_duplicate_customer_workbook_bytes()

    response = client.post(
        "/import/workbook",
        data={
            "year_override": "2026",
            "import_mode": "apply",
            "upsert_policy": "merge",
            "validation_mode": "standard",
            "duplicate_policy": "error",
        },
        files={
            "workbook_file": (
                "planner-duplicate.xlsm",
                workbook_bytes,
                "application/vnd.ms-excel.sheet.macroEnabled.12",
            )
        },
    )

    assert response.status_code == 400
    assert "Import blocked by validation rules" in response.text
    assert "duplicate policy" in response.text.lower()

    with engine.begin() as conn:
        customer_count = conn.execute(
            text("SELECT COUNT(*) FROM customers WHERE cust_code = 'C100'")
        ).scalar_one()
    assert customer_count == 0
